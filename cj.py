import openpyxl
from openpyxl import load_workbook
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from datetime import datetime

root = tk.Tk()
root.title("출장 정보 입력")

entry_frames = []

####################################################################################################################################################################################

def browse_file(entry, is_input_path):
    global 입력경로, 출력경로
    file_path = filedialog.askopenfilename(title="파일 선택")
    entry.delete(0, tk.END)
    entry.insert(0, file_path)
    
    if is_input_path:
        입력경로 = file_path
    else:
        출력경로 = file_path

####################################################################################################################################################################################

def add_file_path_fields():
    file_path_frame = tk.Frame(root)
    file_path_frame.grid(row=0, column=0, pady=5)

    labels = ["파일경로(출장신청목록)", "파일경로(여비상세)", "파일경로(공통항목)"]  # 추가된 부분
    entries = {}

    for i, label in enumerate(labels):
        tk.Label(file_path_frame, text=label).grid(row=i, column=0, padx=5, sticky="e")  # 수정된 부분
        entry = tk.Entry(file_path_frame, width=50)
        browse_button = tk.Button(file_path_frame, text="찾아보기", command=lambda entry=entry, is_input=i<2: browse_file(entry, is_input))  # 수정된 부분
        entry.grid(row=i, column=1, padx=70, sticky="w")  # 수정된 부분
        browse_button.grid(row=i, column=2, padx=5, sticky="w")  # 수정된 부분
        entries[label] = entry

    entry_frames.append((file_path_frame, entries))

####################################################################################################################################################################################

def add_entry_fields():
    entry_frame = tk.Frame(root)
    entry_frame.grid(row=len(entry_frames) + 1, column=0, pady=5)

    labels = ["이름", "직급", "거래처구분(선택)", "생년월일", "입금유형(선택)",  "은행(선택)", "계좌번호(숫자만)", "등급(선택)",
              "출발지(선택)", "도착지(선택)", "교통편(선택)", "정산유형(선택)", "일비(원)"]
    entries = {}

    for i, label in enumerate(labels):
        tk.Label(entry_frame, text=label).grid(row=0, column=i, padx=5, sticky="e")

        if label == "거래처구분(선택)":
            거래처_options = ["10:법인사업자", "20:개인사업자", "30:개인", "40:기타"]
            entry = ttk.Combobox(entry_frame, values=거래처_options, width=15)
        elif label == "입금유형(선택)":
            입금유형_options = ["10:계좌이체", "20:대량이체", "30:원천징수", "40:고지서", "50:CMS", "60:수표", "99:현금"]
            entry = ttk.Combobox(entry_frame, values=입금유형_options, width=15)
        elif label == "은행(선택)":
            entry = ttk.Combobox(entry_frame, width=15)
        elif label == "등급(선택)":
            등급_options = ["1등급", "2등급"]
            entry = ttk.Combobox(entry_frame, values=등급_options, width=15)
        elif label in ["출발지(선택)", "도착지(선택)"]:
            지역_options = ["동두천시청", "직접입력"]
            entry = ttk.Combobox(entry_frame, values=지역_options, width=15)
        elif label == "교통편(선택)":
            교통편_options = ["001:없음","002:자가","003:버스","004:철도","005:항공","006:선박"]
            entry = ttk.Combobox(entry_frame, values=교통편_options, width=15)
        elif label == "정산유형(선택)":
            정산유형_options = ["001:할인정액","002:실비","003:상한액1/2추가","004:상한액3/10추가"]
            entry = ttk.Combobox(entry_frame, values=정산유형_options, width=15)
        elif label == "계좌번호(숫자만)":
            entry = tk.Entry(entry_frame, width=15)
        else:
            entry = tk.Entry(entry_frame, width=9)

        entry.grid(row=1, column=i, padx=5, sticky="w")
        entries[label] = entry

    entry_frames.append((entry_frame, entries))
    
####################################################################################################################################################################################

def display_entries():
    global 인적데이터, 경로데이터, 입력정보
    '''
    0:이름 1:직급 2:거래처구분(선택) 3:생년월일 4:입금유형(선택) 5:은행(선택) 6:계좌번호(숫자만) 7:등급(선택)
    8:출발지(선택) 9:도착지(선택) 10:교통편(선택) 11:정산유형(선택) 12:일비(원)
    '''

    입력정보 = []
    for frame, entries in entry_frames:
        entry_data = []
        for label, entry in entries.items():
            entry_data.append(entry.get())
        입력정보.append(entry_data)
 
    인적데이터 = []
    경로데이터 = []

    인적데이터 = 입력정보[1:]
    경로데이터 = 입력정보[0][0]

    bank_array(경로데이터[2]) #공통항목

    workbook_path_출장신청 = 경로데이터[0] #출장신청목록
    workbook_path_여비상세 = 경로데이터[1] #여비상세
    data_array = 인적데이터

    input_data_to_excel(workbook_path_출장신청, workbook_path_여비상세, data_array)

####################################################################################################################################################################################

def bank_array(공통항목_path):
    workbook = openpyxl.load_workbook(공통항목_path)
    sheet_은행코드 = workbook['은행코드_참고자료']
    
    global 은행_options 
    은행_options = []
    
    for row_num in range(2, 255):
        은행_options.append(sheet_은행코드.cell(row=row_num, column=1).value)

####################################################################################################################################################################################

def 검증():
    print("\n-\n")
    print("입력정보:\n", 입력정보)
    print("\n-\n")
    print("경로데이터:\n", 경로데이터) #[출장신청목록, 여비상세, 공통항목]
    print("\n-\n")
    print("인적데이터:\n", 인적데이터)
    print("\n-\n")
    print("입력경로:\n", 입력경로)
    print("\n-\n")
    print("출력경로:\n", 출력경로)

####################################################################################################################################################################################

def input_data_to_excel(여비신청_path, 여비상세_path, data_array):
     
    # 출장신청목록 정리
    # 엑셀 파일 불러오기
    workbook_출장신청목록 = load_workbook(여비신청_path)

    # 시트 선택
    sheet = workbook_출장신청목록['Col1']

    # 삭제할 행을 저장할 리스트
    rows_to_delete = []

    # 'C' 열을 순회하면서 값이 없는 행 또는 '동두천시...(소속부서)' 찾아서 삭제할 행 리스트에 추가
    for rows in sheet.iter_rows(min_row=5, min_col=2, values_only=True):
        if rows[0] is None or rows[0].split(" ")[0] == '동두천시':
            rows_to_delete.append(rows)

    # 삭제할 행이 있다면 해당 행 삭제
    for row in rows_to_delete:
        sheet.delete_rows(sheet.cell(row=row, column=2).row)

    def has_specific_characters(input_string, characters):
        for char in characters:
            if char in input_string:
                return True
        return False

    # 다른 이름으로 저장
    if has_specific_characters(여비신청_path, "/"):
        split_path = 여비신청_path.split("/")
        출장신청목록정리_path = "/".join(split_path[:-1])
        출장신청목록_modified_path = 출장신청목록정리_path + '/출장신청목록_modified.xlsx'
    else:
        split_path = 여비신청_path.split("\\")
        출장신청목록정리_path = "\\".join(split_path[:-1])
        출장신청목록_modified_path = 출장신청목록정리_path + '/출장신청목록_modified.xlsx'
        
        
    workbook_출장신청목록.save(출장신청목록_modified_path)
    
    workbook_출장신청목록.close()
    '''
    여비상세입력 시트 : A2(1,2) to Y2(25,2)
    여비상세입력설명 시트 : B4(2,4) to Z4(26,4) ...(+1, +2)
    '''
    
    workbook_출장신청목록정리 = load_workbook(출장신청목록_modified_path)
    sheet = workbook_출장신청목록정리['Col1']
    
    # 엑셀 파일 셀
    여비상세_열_start = 1
    여비상세_행_start = 2

    # 출장목록 배열 길이 측정
    column_index = 'B'
    column_data = [cell.value for cell in sheet[column_index]]
    column_length = len(column_data)
   
    # 엑셀 파일 열기
    workbook_여비신청 = openpyxl.load_workbook(출장신청목록_modified_path)
    
    workbook_여비상세 = openpyxl.load_workbook(여비상세_path)

    # 작업 시트 선택
    sheet_출장신청 = workbook_여비신청['Col1']
    
    sheet_여비상세입력 = workbook_여비상세['여비상세입력']
    sheet_여비상세입력설명 = workbook_여비상세['여비상세입력설명']
    
    # 출장신청 데이터 추출
    '''
    B4:순번, C4:구분, D4:출발일자, E4:도착일자, F4:총출장시간, G4:출장지, H4:차량, I4:출장목적, 
    J4:소속, K4:출장자, L4:지출, M4:여비, N4:지출일자, O4:결재상태, P4:여비등급, Q4:비고, R4:업무대행
    '''
    출장시작일_목록=[]
    출장종료일_목록=[]
    일자_목록=[]
    경유지_목록=[] # 출장지
    출장목적_목록=[]
    부서_목록=[]
    실국_목록=[]
    
    for numbs in range(column_length):
        start_numb = 5

        출발일자 = (datetime.strptime((sheet_출장신청['D'+str(start_numb+numbs)].value), '%Y-%m-%d %H:%M')).strftime('%Y%m%d')
        출장시작일_목록.append(출발일자)
        도착일자 = (datetime.strptime((sheet_출장신청['E'+str(start_numb+numbs)].value), '%Y-%m-%d %H:%M')).strftime('%Y%m%d')
        출장종료일_목록.append(도착일자)
        일자_목록.append(도착일자)
        출장지 = sheet_출장신청['G'+str(start_numb+numbs)].value
        경유지_목록.append(출장지)
        출장목적 = sheet_출장신청['I'+str(start_numb+numbs)].value
        출장목적_목록.append(출장목적)
        소속 = sheet_출장신청['J'+str(start_numb+numbs)].value
        실국 = 소속.split(" ")[0]
        실국_목록.append(실국)
        부서 = 소속.split(" ")[1]
        부서_목록.append(부서)
        
    # 데이터 입력
    # data_array = 인적데이터 = [][]
    for rows in range(column_length):
        for numb_of_row in range(len(인적데이터)):
            numb = int(여비상세_행_start) + int(numb_of_row)
            # 여비상세 = A2 = [row][col]
            sheet_여비상세입력.cell(row = numb, column = 여비상세_열_start+0, value = data_array[numb_of_row][0]) #1   A2  B4  "출장자명(100자리이하)" = 0:이름
            sheet_여비상세입력.cell(row = numb, column = 여비상세_열_start+1, value = data_array[numb_of_row][6]) #2   B2  C4  "계좌번호(30자리이하)" = 6:계좌번호(숫자만)
            sheet_여비상세입력.cell(row = numb, column = 여비상세_열_start+2, value = data_array[numb_of_row][1]) #3   C2  D4  "직급명(100자리이하)" = 1:직급
            sheet_여비상세입력.cell(row = numb, column = 여비상세_열_start+3, value = 출장목적_목록[rows]) #4   D2  E4  "출장목적(100자리이하)" = 출장신청파일
            sheet_여비상세입력.cell(row = numb, column = 여비상세_열_start+4, value = data_array[numb_of_row][11]) #5   E2  F4  "정산유형(선택)" = 11:정산유형(선택)
            sheet_여비상세입력.cell(row = numb, column = 여비상세_열_start+5, value = 일자_목록[rows]) #6   F2  G4  "일자(8자리)" = 출장신청파일
            sheet_여비상세입력.cell(row = numb, column = 여비상세_열_start+6, value = data_array[numb_of_row][8]) #7   G2  H4  "출발지(100자리이하)" = 8:출발지(선택)
            sheet_여비상세입력.cell(row = numb, column = 여비상세_열_start+7, value = 경유지_목록[rows]) #8   H2  I4  "경유지(100자리이하)" = 출장신청파일
            sheet_여비상세입력.cell(row = numb, column = 여비상세_열_start+8, value = data_array[numb_of_row][9]) #9   I2  J4  "도착지(100자리이하)" = 9:도착지(선택)
            sheet_여비상세입력.cell(row = numb, column = 여비상세_열_start+9, value = data_array[numb_of_row][10]) #10  J2  K4  "교통편(선택)" = 10:교통편(선택)
            sheet_여비상세입력.cell(row = numb, column = 여비상세_열_start+10, value = "") #11  K2  L4  "종별(100자리이하)" = ""
            sheet_여비상세입력.cell(row = numb, column = 여비상세_열_start+11, value = "") #12  L2  M4  "등급(100자리이하)" = ""
            sheet_여비상세입력.cell(row = numb, column = 여비상세_열_start+12, value = "") #13  M2  N4  "거리(10자리이하)" = ""
            sheet_여비상세입력.cell(row = numb, column = 여비상세_열_start+13, value = "") #14  N2  O4  "요금(17자리이하)" = ""
            sheet_여비상세입력.cell(row = numb, column = 여비상세_열_start+14, value = 출장시작일_목록[rows]) #15  O2  P4  "출장시작일(8자리)" = 출장신청파일
            sheet_여비상세입력.cell(row = numb, column = 여비상세_열_start+15, value = 출장종료일_목록[rows]) #16  P2  Q4  "출장종료일(8자리)" = 출장신청파일
            sheet_여비상세입력.cell(row = numb, column = 여비상세_열_start+16, value = "") #17  Q2  R4  "식비(17자리이하)" = ""
            sheet_여비상세입력.cell(row = numb, column = 여비상세_열_start+17, value = "") #18  R2  S4  "숙박료(17자리이하)" = ""
            sheet_여비상세입력.cell(row = numb, column = 여비상세_열_start+18, value = data_array[numb_of_row][12]) #19  S2  T4  "일비(17자리이하)" = 12:일비(원)
            sheet_여비상세입력.cell(row = numb, column = 여비상세_열_start+19, value = "") #20  T2  U4  "현지교통비(사용하지 않는 항목)" = ""
            sheet_여비상세입력.cell(row = numb, column = 여비상세_열_start+20, value = "") #21  U2  V4  "기타(사용하지 않는 항목)" = ""
            sheet_여비상세입력.cell(row = numb, column = 여비상세_열_start+21, value = data_array[numb_of_row][12]) #22  V2  W4  "계(17자리이하)" = 12:일비(원)
            sheet_여비상세입력.cell(row = numb, column = 여비상세_열_start+22, value = data_array[numb_of_row][12]) #23  W2  X4  "청구 및 수령액(17자리이하)" = 12:일비(원)
            sheet_여비상세입력.cell(row = numb, column = 여비상세_열_start+23, value = 실국_목록[rows]) #24  X2  Y4  "실국명(100자리이하)" = 출장신청파일
            sheet_여비상세입력.cell(row = numb, column = 여비상세_열_start+24, value = 부서_목록[rows]) #25  Y2  Z4  "부서명(100자리이하)" = 출장신청파일
            # 입력설명 = B4 = [row+2][col+1]
            sheet_여비상세입력설명.cell(row = numb+2, column = 여비상세_열_start+0+1, value = data_array[numb_of_row][0]) #1   A2  B4  "출장자명(100자리이하)" = 0:이름
            sheet_여비상세입력설명.cell(row = numb+2, column = 여비상세_열_start+1+1, value = data_array[numb_of_row][6]) #2   B2  C4  "계좌번호(30자리이하)" = 6:계좌번호(숫자만)
            sheet_여비상세입력설명.cell(row = numb+2, column = 여비상세_열_start+2+1, value = data_array[numb_of_row][1]) #3   C2  D4  "직급명(100자리이하)" = 1:직급
            sheet_여비상세입력설명.cell(row = numb+2, column = 여비상세_열_start+3+1, value = 출장목적_목록[rows]) #4   D2  E4  "출장목적(100자리이하)" = 출장신청파일
            sheet_여비상세입력설명.cell(row = numb+2, column = 여비상세_열_start+4+1, value = data_array[numb_of_row][11]) #5   E2  F4  "정산유형(선택)" = 11:정산유형(선택)
            sheet_여비상세입력설명.cell(row = numb+2, column = 여비상세_열_start+5+1, value = 일자_목록[rows]) #6   F2  G4  "일자(8자리)" = 출장신청파일
            sheet_여비상세입력설명.cell(row = numb+2, column = 여비상세_열_start+6+1, value = data_array[numb_of_row][8]) #7   G2  H4  "출발지(100자리이하)" = 8:출발지(선택)
            sheet_여비상세입력설명.cell(row = numb+2, column = 여비상세_열_start+7+1, value = 경유지_목록[rows]) #8   H2  I4  "경유지(100자리이하)" = 출장신청파일
            sheet_여비상세입력설명.cell(row = numb+2, column = 여비상세_열_start+8+1, value = data_array[numb_of_row][9]) #9   I2  J4  "도착지(100자리이하)" = 9:도착지(선택)
            sheet_여비상세입력설명.cell(row = numb+2, column = 여비상세_열_start+9+1, value = data_array[numb_of_row][10]) #10  J2  K4  "교통편(선택)" = 10:교통편(선택)
            sheet_여비상세입력설명.cell(row = numb+2, column = 여비상세_열_start+10+1, value = "") #11  K2  L4  "종별(100자리이하)" = ""
            sheet_여비상세입력설명.cell(row = numb+2, column = 여비상세_열_start+11+1, value = "") #12  L2  M4  "등급(100자리이하)" = ""
            sheet_여비상세입력설명.cell(row = numb+2, column = 여비상세_열_start+12+1, value = "") #13  M2  N4  "거리(10자리이하)" = ""
            sheet_여비상세입력설명.cell(row = numb+2, column = 여비상세_열_start+13+1, value = "") #14  N2  O4  "요금(17자리이하)" = ""
            sheet_여비상세입력설명.cell(row = numb+2, column = 여비상세_열_start+14+1, value = 출장시작일_목록[rows]) #15  O2  P4  "출장시작일(8자리)" = 출장신청파일
            sheet_여비상세입력설명.cell(row = numb+2, column = 여비상세_열_start+15+1, value = 출장종료일_목록[rows]) #16  P2  Q4  "출장종료일(8자리)" = 출장신청파일
            sheet_여비상세입력설명.cell(row = numb+2, column = 여비상세_열_start+16+1, value = "") #17  Q2  R4  "식비(17자리이하)" = ""
            sheet_여비상세입력설명.cell(row = numb+2, column = 여비상세_열_start+17+1, value = "") #18  R2  S4  "숙박료(17자리이하)" = ""
            sheet_여비상세입력설명.cell(row = numb+2, column = 여비상세_열_start+18+1, value = data_array[numb_of_row][12]) #19  S2  T4  "일비(17자리이하)" = 12:일비(원)
            sheet_여비상세입력설명.cell(row = numb+2, column = 여비상세_열_start+19+1, value = "") #20  T2  U4  "현지교통비(사용하지 않는 항목)" = ""
            sheet_여비상세입력설명.cell(row = numb+2, column = 여비상세_열_start+20+1, value = "") #21  U2  V4  "기타(사용하지 않는 항목)" = ""
            sheet_여비상세입력설명.cell(row = numb+2, column = 여비상세_열_start+21+1, value = data_array[numb_of_row][12]) #22  V2  W4  "계(17자리이하)" = 12:일비(원)
            sheet_여비상세입력설명.cell(row = numb+2, column = 여비상세_열_start+22+1, value = data_array[numb_of_row][12]) #23  W2  X4  "청구 및 수령액(17자리이하)" = 12:일비(원)
            sheet_여비상세입력설명.cell(row = numb+2, column = 여비상세_열_start+23+1, value = 실국_목록[rows]) #24  X2  Y4  "실국명(100자리이하)" = 출장신청파일
            sheet_여비상세입력설명.cell(row = numb+2, column = 여비상세_열_start+24+1, value = 부서_목록[rows]) #25  Y2  Z4  "부서명(100자리이하)" = 출장신청파일            

    # 변경사항 저장
    workbook_여비상세.save(여비상세_path)
    workbook_여비상세.close()
    print("데이터가 엑셀에 입력되었습니다.")

####################################################################################################################################################################################

add_file_path_fields()

add_button = tk.Button(root, text="입력 추가", command=add_entry_fields)
add_button.grid(row=1, column=0, pady=10)

display_button = tk.Button(root, text="입력완료", command=display_entries)
display_button.grid(row=len(entry_frames) + 20, column=0, pady=10)

display_button = tk.Button(root, text="데이터검증", command=검증)
display_button.grid(row=len(entry_frames) + 40, column=0, pady=10)

root.mainloop()
